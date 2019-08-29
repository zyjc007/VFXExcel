#	-*- coding:UTF-8 -*-

import openpyxl
import argparse
import os

#	获取命令行参数
def getCommandLine():
	parser = argparse.ArgumentParser(description = 'changename')
	parser.add_argument('Excel_path', metavar = 'EXCEL_PATH', type = str, nargs = 1, help = 'path of the Excel file')
	parser.add_argument('List_path', metavar = 'LIST_PATH', type = str, nargs = 1, help = 'path of the List file')

	return parser

#	获取文件路径
def getPath():
	parser = getCommandLine()
	args = vars(parser.parse_args())
	excelPath = args['Excel_path'][0]
	listPath = args['List_path'][0]
	
	ParseExcel(excelPath, listPath)

#	解析Excel
def ParseExcel(excelPath, listPath):
	#	获取表格对象
	wb = openpyxl.load_workbook(excelPath)
	#	获取表单对象
	ws = wb.active

	#	获取第几集
	oEP = ws['B']
	#	获取第几刀
	clipnum = ws['C']
	#	获取是否多层
	level = ws['F']

	nameList = []
	for i in range(0,len(oEP)):
		if str(level[i].value) == 'None':
			name = str(oEP[i].value) + '_VFX_' + str(clipnum[i].value)
		else:
			name = str(oEP[i].value) + '_VFX_' + str(clipnum[i].value) + '_' + str(level[i].value)

		nameList.append(name) 

	nameList.pop(0)

	ParseList(nameList, listPath)

def ParseList(newNameList, listPath):

	filename = os.listdir(listPath)
	filename.remove('.DS_Store')
	filename.sort()
	
	print(newNameList)
	print(filename)

	index = 0
	for i in filename:
		oldname = filename[index]
		newname = str(newNameList[index]) + '.mov'
		print(oldname, '=====>', newname)
		os.rename(os.path.join(listPath,oldname), os.path.join(listPath, newname))
		index = index + 1


if __name__ == '__main__':
	getPath()